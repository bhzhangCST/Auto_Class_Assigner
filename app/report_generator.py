"""
Report Generator Module
Generates multi-sheet Excel output with class assignment results.
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def generate_result_excel(df: pd.DataFrame, special_df: Optional[pd.DataFrame],
                          subject_cols: List[str],
                          output_path: Path, grade_name: str) -> Path:
    """Generate class assignment result Excel file with per-class sheets."""
    output_file = output_path / f"{grade_name}年级分班结果.xlsx"

    output_cols = ['姓名', '原班级', '总分'] + subject_cols + ['年级排名']
    available_cols = [col for col in output_cols if col in df.columns]

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center', vertical='center')

    classes = sorted(df['新班级'].unique(), key=lambda x: int(x.replace('班', '')))

    grade_avg_total = df['总分'].mean()
    grade_avg_subjects = {col: df[col].mean() for col in subject_cols}

    # Per-class sheets
    for class_name in classes:
        class_df = df[df['新班级'] == class_name][available_cols].copy()
        class_df = class_df.sort_values('年级排名')

        ws = wb.create_sheet(title=class_name)

        for r_idx, row in enumerate(dataframe_to_rows(class_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = center
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        for c in range(1, len(available_cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12

        # Class statistics at bottom
        sr = len(class_df) + 3
        ws.cell(row=sr, column=1, value="班级统计").font = Font(bold=True)

        ws.cell(row=sr + 1, column=1, value="总分平均")
        ws.cell(row=sr + 1, column=2, value=round(class_df['总分'].mean(), 2))
        ws.cell(row=sr + 1, column=3, value=f"(年级平均: {round(grade_avg_total, 2)})")

        for i, col in enumerate(subject_cols):
            ws.cell(row=sr + 2 + i, column=1, value=f"{col}平均")
            ws.cell(row=sr + 2 + i, column=2, value=round(class_df[col].mean(), 2))
            ws.cell(row=sr + 2 + i, column=3, value=f"(年级平均: {round(grade_avg_subjects[col], 2)})")

    # Summary sheet
    ws_summary = wb.create_sheet(title="统计汇总")
    summary_headers = ['班级', '人数', '总分平均'] + [f'{col}平均' for col in subject_cols]

    for c_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for r_idx, class_name in enumerate(classes, 2):
        class_df = df[df['新班级'] == class_name]
        ws_summary.cell(row=r_idx, column=1, value=class_name).alignment = center
        ws_summary.cell(row=r_idx, column=2, value=len(class_df)).alignment = center
        ws_summary.cell(row=r_idx, column=3, value=round(class_df['总分'].mean(), 2)).alignment = center
        for i, col in enumerate(subject_cols):
            ws_summary.cell(row=r_idx, column=4 + i, value=round(class_df[col].mean(), 2)).alignment = center

    gr = len(classes) + 2
    ws_summary.cell(row=gr, column=1, value="年级平均").font = Font(bold=True)
    ws_summary.cell(row=gr, column=2, value=len(df))
    ws_summary.cell(row=gr, column=3, value=round(grade_avg_total, 2))
    for i, col in enumerate(subject_cols):
        ws_summary.cell(row=gr, column=4 + i, value=round(grade_avg_subjects[col], 2))

    for c in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(c)].width = 12

    # Special students sheet (absent, long-term leave, etc.)
    if special_df is not None and len(special_df) > 0:
        ws_special = wb.create_sheet(title="特殊情况")
        special_cols = ['姓名', '原班级'] + subject_cols
        available_special = [col for col in special_cols if col in special_df.columns]

        for c_idx, header in enumerate(available_special, 1):
            cell = ws_special.cell(row=1, column=c_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E57373", end_color="E57373", fill_type="solid")
            cell.alignment = center

        for r_idx, (_, row) in enumerate(special_df.iterrows(), 2):
            for c_idx, col in enumerate(available_special, 1):
                cell = ws_special.cell(row=r_idx, column=c_idx, value=row.get(col, ''))
                cell.alignment = center

        for c in range(1, len(available_special) + 1):
            ws_special.column_dimensions[get_column_letter(c)].width = 12

    wb.save(output_file)
    return output_file


def grade_number_to_chinese(grade: str) -> str:
    """Convert Arabic numeral grade to Chinese."""
    mapping = {'1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六'}
    return mapping.get(grade, grade)
